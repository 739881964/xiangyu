from openpyxl import *
from scripts.get_cfg import config
from scripts import base_path


class ExcelClass(object):

    def __init__(self, path, sheet_name):
        self.path = path
        self.sheet_name = sheet_name

    def read_excel_all_data(self):
        wb = load_workbook(self.path)
        sheet = wb[self.sheet_name]

        head_list_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        other_list_data = []
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            all_data = dict(zip(head_list_data, data))
            other_list_data.append(all_data)

        return other_list_data

    def write_data_in_excel(self, row, actual, res):
        wb = load_workbook(self.path)
        if self.sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]

        if isinstance(row, int) and (2 <= row <= sheet.max_row):
            sheet.cell(row, config.get_int('excel', 'actual_col')).value = actual
            sheet.cell(row, config.get_int('excel', 'res_col')).value = res
            wb.save(self.path)

        else:
            print('行号要大于1')


if __name__ == '__main__':
    _path = base_path.TEST_DATAS_EXCEL_PATH
    _sheet_name = '加法'
    excel = ExcelClass(_path, _sheet_name)
    print(excel.read_excel_all_data())
