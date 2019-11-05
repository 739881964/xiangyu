# !/usr/bin/env python
# -*- coding:utf-8 -*-
# author:xiangyu
# phone：19942429056
# datetime:2019/9/4 10:45
# software: PyCharm
import os

from openpyxl import *
from scripts.get_cfg import config
from scripts import base_path


class ExcelClass(object):

    def __init__(self, path, sheet_name):
        self.path = path
        self.sheet_name = sheet_name

    def read_excel_all_data(self):
        wb = load_workbook(self.path)
        sheet = wb[self.sheet_name]

        head_list_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        other_list_data = []
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            all_data = dict(zip(head_list_data, data))
            other_list_data.append(all_data)

        return other_list_data

    def write_data_in_excel(self, row, actual, res):
        wb = load_workbook(self.path)
        if self.sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]

        if isinstance(row, int) and (2 <= row <= sheet.max_row):
            sheet.cell(row, config.get_int('excel', 'actual_col')).value = actual
            sheet.cell(row, config.get_int('excel', 'res_col')).value = res
            wb.save(self.path)
        else:
            # pass
            print('行号要大于1')


if __name__ == '__main__':
    pass
    # _path = r'C:\Users\xiangyu\Desktop\人工识别统计模板.xlsx'
    # _sheet_name = 'Sheet1'
    # excel = ExcelClass(_path, _sheet_name)
    # # print(excel.read_excel_all_data())
    # path = r'C:\Users\xiangyu\Desktop\new_word_file\duolian_word_list_file'
    # files = os.listdir(path)
    # command = list(map(lambda x: x.split('.')[0], files))
    # for i in range(len(command)):
    #     excel.write_data_in_excel(i + 4, command[i])
    # pass
