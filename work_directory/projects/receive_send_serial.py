# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/11/12 11:31
# @Author  : Xiang Yu
# @File    : receive_send_serial.py
# @Software: PyCharm
# @Company : BEIJING INTENGINE


__all__ = ["PandasManual", "OperationSerial", 'main']


import sys
import time

import serial
import serial.tools.list_ports
from openpyxl import load_workbook


class PandasManual:
    """ openpyxl operate excel """

    def __init__(self, file_path, sheet):
        # super().__init__(file_path)
        self.file_path = file_path
        self.sheet = sheet

    def read_data(self, sheet='') -> list:
        """
        read excel file by openpyxl
        return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = wb[self.sheet]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = list()
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list


class OperationSerial:
    """ 实现串口与主控MCU之间的通信：数据接收与发送 """

    def __init__(self, port='', is_debug=False):
        """ 初始化连接串口，接收端口和波特率，一般为：9600 或 115200 """
        # self.port = port
        if port:
            if is_debug:
                self.port = serial.Serial(port, 115200)
            else:
                self.port = serial.Serial(port, 9600)
            if not self.port:
                self.port.open()

    def is_open(self):
        """ 打开串口 """
        if not self.port.is_open:
            self.port.open()

    def close_port(self):
        """ 关闭串口 """
        self.port.close()

    @classmethod
    def get_port_list(cls):
        # 获取windows所有端口
        port_list = serial.tools.list_ports.comports()
        all_port = []
        for port in port_list:
            all_port.append(port[0])

        return all_port

    def debug(self):
        """
        接收串口数据 debug 模式 115200 波特率
        :return: None
        """
        while True:
            data = str(self.port.readline().decode('utf8'))
            # data = str(self.port.read(byte).decode('utf8'))  # decode('utf8'))
            # data = str(self.port.readline(), 'utf8')
            # with open(self._path, 'a+') as f:
            #     f.write(data + '\n')
            # if 'SpottingWordList' in data:
            print(data)

    def receive(self, byte: int, excel_data: list):
        """
        接收串口数据 串口通讯模式 9600 波特率
        :param byte: 接收的字节数
        :param excel_data: excel读取的嵌套字典数据
        :return: None
        """
        while True:
            data = self.port.read(byte)
            # print(data)
            # if len(data) == 13:
            if data:
                r = list(map(hex, list(data)))
                # print(r)
                # ret = r[-3:] + r[:-3]
                res = ' '.join(list(map(lambda y: y.replace('x', '').upper()[-2:], r)))
                for excel in excel_data:
                    if excel['语音模块发送给主控MCU的UART数据'] == res:
                        print('响应串口码: [{}] -> 对应的命令或状态：[{}]'.format(res, excel['命令 或 状态']))
                        break
                else:
                    print('响应的串口码 [{}] 无法在excel中找到'.format(res))
                # print(res)
                print()

    def send(self, serial_data: list):
        """
        发送数据给串口 串口通讯模式 9600 波特率
        :param serial_data: 发送的串口数据
        :return: None
        """
        # _data = ' '.join([hex(ord(c)).replace('0x', '') for c in data])
        for i in range(len(serial_data)):
            one_data = serial_data[i]['主控MCU发送给语音模块的UART数据']
            _data = bytes.fromhex(one_data)
            try:
                self.port.write(_data)
                print('命令或状态: [{}] 串口码：[{}] -> 发送成功 '.format(serial_data[i]['命令 或 状态'], one_data))
                while True:
                    result = input('请输入语音反馈是否正确(y/n/q?)： ')
                    try:
                        if (result == 'y') or (result == 'Y'):
                            print('发送串口码 [{}] -> MCU -> 语音响应成功'.format(one_data))
                            break
                        elif (result == 'n') or (result == 'N'):
                            print('发送串口码 [{}] -> MCU -> 语音响应错误'.format(one_data))
                            break
                        elif (result == 'q') or (result == 'Q'):
                            return
                    except:
                        print('请输入合法的测试结果(y/n/q?)')
            except:
                print('命令或状态: [{}] 串口码：[{}] -> 发送失败 '.format(serial_data[i]['命令 或 状态'], one_data))
            print()

    def receive_and_send(self, byte: int, dic_data: list, sleep_time: int):
        """
        接收串口数据并点发送excel中相应的串口数据
        :param byte: 接收串口字节数
        :param dic_data: 读取的excel数据
        :param sleep_time: 识别与发送串口码的间隔时间
        :return: None
        """
        while True:
            data = self.port.read(byte)
            if data:
                r = list(map(hex, list(data)))
                # print(r)
                # ret = r[-3:] + r[:-3]
                res = ' '.join(list(map(lambda y: y.replace('x', '').upper()[-2:], r)))
                for i in range(len(dic_data)):
                    if dic_data[i]['语音模块发送给主控MCU的UART数据'] == res:
                        print('响应的串口码: [{}] -> 命令或状态：[{}]'.format(res, dic_data[i]['命令 或 状态']))
                        time.sleep(sleep_time)
                        one_data = dic_data[i]['主控MCU发送给语音模块的UART数据']
                        _data = bytes.fromhex(one_data)
                        try:
                            self.port.write(_data)
                            print('命令或状态: [{}] -> 发给主控MCU串口码：[{}] -> 发送成功 '.format(dic_data[i]['命令 或 状态'], one_data))
                            while True:
                                result = input('请输入语音反馈是否正确(y/n/q?)： ')
                                try:
                                    if result == 'y':
                                        print('发送串口码 [{}] -> MCU -> 语音响应成功'.format(one_data))
                                        break
                                    elif result == 'n':
                                        print('发送串口码 [{}] -> MCU -> 语音响应错误'.format(one_data))
                                        break
                                except:
                                    print('请输入合法的测试结果(y/n/q?)')
                        except:
                            print('命令或状态: [{}] 串口码：[{}] -> 发送失败 '.format(dic_data[i]['命令 或 状态'], one_data))
                        break
                else:
                    print('响应的串口码 [{}] 无法在excel中找到'.format(res))
                print()


def main(r_data, port, model, recv_bytes, sp_time):
    """
    根据需要选择测试的模式
    :param r_data: excel预期接收到的串口码
    :param port: 设备连接的端口号
    :param model: 测试模式
    :param recv_bytes: 接收串口数据字节数
    :param sp_time: 接收串口数据与发送串口数据之间的间隔时间
    :return: None
    """
    if model == 'SPD':
        # debug模式
        obj = OperationSerial(port, is_debug=True)
        obj.debug()
    elif model == 'SPR':
        # 串口通讯接收模式
        obj = OperationSerial(port)
        obj.receive(recv_bytes, r_data)
    elif model == 'SPS':
        # 串口通讯发送模式
        obj = OperationSerial(port)
        obj.send(r_data)
    elif model == 'SPT':
        # 串口通讯接收与发送模式
        obj = OperationSerial(port)
        obj.receive_and_send(recv_bytes, r_data, sp_time)


def run():
    try:
        panda = PandasManual(r'C:\Users\xiangyu\Desktop\串口数据.xlsx', sheet='串口通信数据详表以及语音播报内容')
        all_data = panda.read_data()
    except Exception as e:
        print(e)
        sys.exit()
    else:
        # 根据项目需求：输入端口、模式、接收字节数、接收与发送数据的时间间隔参数
        main(all_data, port='COM4', model='SPD', recv_bytes=13, sp_time=3)
    finally:
        print()
        print('Test End !')


if __name__ == "__main__":
    # sp = OperationSerial()
    # print(sp.get_port_list())
    run()

    pass
