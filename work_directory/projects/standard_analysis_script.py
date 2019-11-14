# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/10/10 14:51
# @Author  : Xiang Yu
# @File    : standard_analysis_script.py
# @Software: PyCharm
# @Company : BEIJING INTENGINE


"""
识别率分析脚本  分析日志应放在D盘根目录下，只需要两个.log文件即可(MIC和slaver_board开头)，
其他的文件名称含有MIC或slaver_board的建议删除，唤醒词的个数根据项目需要修改，修改 640行 的 __n 参数即可。
生成的测试结果在D盘根目录下为 "test_result.xlsx" 的 excel文件
"""


__all__ = [
        'FixExcel', 'PandasManual', 'get_res_count_data', 'get_every_command_times',
        'get_all_command_times', 'run_time', 'get_start_time_list', 'get_all_broadcast_wav',
        'read_rs_trip_data', 'get_slaver_board_log', 'get_mic_log', 'get_lost_wav_start_time',
        'get_lost_wav', 'get_lost_command', 'operation', 'recognize_rate', 'main'
]


import sys
import os
import re
import time

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from warnings import simplefilter

simplefilter(action='ignore', category=FutureWarning)


class FixExcel:

    def __init__(self, file_path):
        self.file_path = file_path

    def modify_excel(self) -> None:
        wb = load_workbook(self.file_path)
        sheet_names = wb.sheetnames

        for k in range(len(sheet_names)):
            one_sheet = wb[sheet_names[k]]
            # 水平居中，垂直居中
            # one_sheet.alignment = Alignment(horizontal='center', vertical='center')
            # 获取每一列的内容的最大宽度
            m = 0
            # 每列
            col_width = list()
            for col in one_sheet.columns:
                # 每行
                for j in range(len(col)):
                    if j == 0:
                        # list增加一个元素
                        col_width.append(len(str(col[j].value)))
                    else:
                        # 获得每列中的内容的最大宽度
                        if col_width[m] < len(str(col[j].value)):
                            col_width[m] = len(str(col[j].value))
                m = m + 1

            # 设置列宽
            for m in range(len(col_width)):
                # 根据列的数字返回字母
                col_letter = get_column_letter(m + 1)
                # 当宽度大于80，宽度设置为80
                if col_width[m] > 80:
                    one_sheet.column_dimensions[col_letter].width = 80
                # 只有当宽度大于4，才设置列宽
                elif col_width[m] > 4:
                    one_sheet.column_dimensions[col_letter].width = col_width[m] + 4

        wb.save(filename=self.file_path)


class PandasManual(FixExcel):
    """ pandas operate excel """

    def __new__(cls, *args, **kwargs):
        """ only create one object address """
        if not hasattr(PandasManual, '_instance'):
            cls._instance = super(PandasManual, cls).__new__(cls)
            # print(cls._instance)  obj -> address
            return cls._instance

    def __init__(self, file_path):
        super().__init__(file_path)
        # self.file_path = file_path
        self.sheet_name = []

    def read_csv(self, sheet=None, encoding='gbk'):
        """ read csv file content """
        data = None
        try:
            if sheet:
                data = pd.read_csv(self.file_path, sheet_name=sheet, encoding=encoding)
            else:
                data = pd.read_csv(self.file_path, encoding=encoding)
        except Exception as e:
            print(e)

        return data

    def get_data(self, sheet=None):
        """ read excel whole data by pandas """
        if sheet:
            data = pd.read_excel(self.file_path, sheet_name=sheet)
        else:
            data = pd.read_excel(self.file_path, sheet_name=self.sheet_name[1])

        return data

    def read_data(self, sheet=None):
        """
        read excel file by openpyxl
        return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = wb[self.sheet_name[1]]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = list()
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list

    # @run_time()
    def write_data(self, rows, columns, sheet=None, *args):
        """
        write in excel depend on row by openpyxl
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = Workbook().create_sheet('pass_fail_info')

        for row in rows:
            if isinstance(row, int) and 2 <= row <= sheet.max_row:
                for col in columns:
                    try:
                        sheet.cell(row, col).value = args
                        wb.save(self.file_path)
                    except Exception as e:
                        print(e)
            else:
                print('row is not exist')

    # @run_time()
    def excel_add_sheet(self, data: list, sheet: list, _sheet_name='首页', header=None, index=None):
        """ not kill before operation when write after depend on column by pandas """
        ''' if not exists excel result, then create it '''
        try:
            if not os.path.exists(self.file_path):
                df = pd.DataFrame(['测试结果已生成!'])
                df.to_excel(self.file_path, sheet_name=_sheet_name, header=header, index=index)
                print('创建测试结果文件xlsx成功！')
        except Exception as e:
            print(f'xlsx测试结果文件已存在，自动创建失败-{e}')
            pass

        with pd.ExcelWriter(self.file_path) as excel_writer:
            book = load_workbook(excel_writer.path)
            excel_writer.book = book

            for i in range(len(data)):
                self.sheet_name.append(sheet[i])
                df = pd.DataFrame(data[i])
                df.to_excel(excel_writer=excel_writer, sheet_name=sheet[i], index=False)

            excel_writer.save()
        # excel_writer.close()


def get_res_count_data(data) -> list:
    """ get Chinese depend on data where from slaver_board.log """
    _count_data = list()
    pattern = re.compile('[\u4e00-\u9fa5]+')
    for i in data:  # i is str
        # __data = re.findall('[\u4e00-\u9fa5]+', i)
        if pattern.findall(i):
            _count_data.append(i)

    return _count_data


def get_every_command_times(data) -> list:
    """ get Chinese depend on data where from slaver_board.log """
    count_data = list()
    pattern = re.compile('[\u4e00-\u9fa5]+')
    for i in data:  # i is str
        res = pattern.findall(i)
        if res and 'NO' not in i:
            count_data.append(res[0])

    return count_data


def get_all_command_times(a_list) -> tuple:
    """ 统计每个命令次出现的次数-list """
    one_list = list()
    for i in a_list:
        if i not in one_list:
            one_list.append(i)

    times = list(map(a_list.count, one_list))

    return one_list, times


def run_time(num: int = 0):
    """ calculate test run time log have time list """
    def count_run_time(func):
        def run(*args, **kwargs):
            one_time = time.perf_counter()
            res = func(*args, **kwargs)
            spent_time = time.perf_counter() - one_time
            if spent_time > num:
                print('测试执行的时间为: %.2f 秒' % spent_time)
            return res
        return run
    return count_run_time


def get_start_time_list(mic_data: '播放日志', slavery_data: '识别日志' = None) -> list:
    """ log no time list """
    date_time = []
    for i in list(map(lambda x: x[1:11], slavery_data)):
        if i not in date_time:
            date_time.append(i)
    start_time_list = list()
    # pattern = re.compile('\*\*\*(\w+)\*\*')
    start = mic_data[0].split()[-1]
    for i in mic_data:
        if '***' in i:
            in_time = i[5:-2]
            if in_time >= start:
                whole_time = ' '.join([date_time[0], in_time])
                start_time_list.append(whole_time)
            else:
                whole_time = ' '.join([date_time[1], in_time])
                start_time_list.append(whole_time)

    return start_time_list


def get_all_broadcast_wav(data) -> list:
    """ 获取全部播放的音频 """
    wav_list = list()
    # pattern = re.compile(r'\\\\.*\.wav')
    pattern = re.compile('.*\.wav')
    for i in data:
        if pattern.findall(i):
            res = pattern.findall(i)[0]
            # resp = 'D' + res + 'wav'
            wav_list.append(res)

    return wav_list


def read_rs_trip_data(file_name) -> list:
    """ abandon \n from data to list """
    try:
        with open(file_name, 'r', encoding='gbk') as f:
            data = list(map(lambda line: line.rstrip('\n'), f.readlines()))
    except:
        with open(file_name, 'r', encoding='utf8') as f:
            data = list(map(lambda line: line.rstrip('\n'), f.readlines()))

    return data


def get_slaver_board_log(base_path) -> str:
    """ 获取识别到语音的 slaver.log """
    d_files = os.listdir(base_path)
    for file in d_files:
        # if ('slaver_board' and '.log') in file:
        if 'slaver_board' in file:
            slaver_board_file = os.path.join(base_path, file)

            return slaver_board_file


def get_mic_log(base_path) -> str:
    """ 获取播放语音的 MIC.log """
    d_files = os.listdir(base_path)
    for file in d_files:
        # if ('MIC' and '.log') in file:
        if 'MIC' in file:
            mic_file = os.path.join(base_path, file)

            return mic_file


def get_lost_wav_start_time(all_start_time, data) -> list:
    """ 获取未识别语音播放开始时间 """
    one_data = data['start_time'].tolist()
    time_list = list(filter(lambda x: x not in one_data, all_start_time))
    # time_list = []
    # for time in all_start_time:
    #     if time not in one_data:
    #         time_list.append(time)

    return time_list


def get_lost_wav(require_wav, data) -> list:
    """
    获取未识别的音频
    data: 必须是 DataFrame 形式的数据，用 pandas.DataFrame() 转换
    """
    res = data['wav_name'].tolist()
    # lost_wav = list(filter(lambda x: True if x not in res else False, require_wav))
    lost_wav = list(filter(lambda x: x not in res, require_wav))
    # lost_wav = []
    # for wav in require_wav:
    #     if wav not in res:
    #         lost_wav.append(wav)

    return lost_wav


def get_lost_command(all_wav, require_wav, need_command) -> list:
    """ 通过音频名称获取未识别的命令词 """
    commands = list()
    for i in range(len(all_wav)):
        for j in range(len(require_wav)):
            if all_wav[i] == require_wav[j]:
                commands.append(need_command[j])
                break

    return commands


def operation(log_res_len, test_count, need_wav, compare_command) -> tuple:
    """ 计算 pass 和 fail """
    num = 1
    case_id = list()
    wav_name = list()
    start_time = list()
    expected_command = list()
    reback_time = list()
    reback_command = list()
    signel = list()
    for i in range(len(test_count)):
        for j in range(len(log_res_len)):
            key = log_res_len[j].split()[-1]  # 命令词
            one_time = log_res_len[j][1:20]  # 时间所在的切片位置
            try:
                if test_count[i+1] > one_time > test_count[i]:
                    if key == compare_command[i]:
                        case_id.append(num)
                        wav_name.append(need_wav[i])
                        start_time.append(test_count[i])
                        expected_command.append(compare_command[i])
                        reback_time.append(one_time)
                        reback_command.append(key)
                        signel.append('pass')
                        num += 1
                        break
                    else:
                        case_id.append(num)
                        wav_name.append(need_wav[i])
                        start_time.append(test_count[i])
                        expected_command.append(compare_command[i])
                        reback_time.append(one_time)
                        reback_command.append(key)
                        signel.append('fail')
                        num += 1
                        break
            except:
                if i == len(test_count) - 1:
                    if one_time > test_count[i]:
                        if key == compare_command[i]:
                            case_id.append(num)
                            wav_name.append(need_wav[i])
                            start_time.append(test_count[i])
                            expected_command.append(compare_command[i])
                            reback_time.append(one_time)
                            reback_command.append(key)
                            signel.append('pass')
                            num += 1
                            break
                        else:
                            case_id.append(num)
                            wav_name.append(need_wav[i])
                            start_time.append(test_count[i])
                            expected_command.append(compare_command[i])
                            reback_time.append(one_time)
                            reback_command.append(key)
                            signel.append('fail')
                            num += 1
                            break

    return case_id, wav_name, start_time, expected_command, reback_time, reback_command, signel


def recognize_rate(data, count_times, times, awake_command, time_list, all_command, count_command) -> tuple:

    """ 计算识别率 """

    # 正确唤醒词
    ak_right = list(filter(lambda y: y['signel'] == 'pass',
                           list(filter(lambda x: x['expected_command'] in awake_command, data))))
    # 错误唤醒词
    ak_error = list(filter(lambda y: y['signel'] == 'fail',
                           list(filter(lambda x: x['expected_command'] in awake_command, data))))

    # 唤醒词未识别次数
    ak_null = times - len(ak_error) - len(ak_right)

    # 其他命令词的正确率
    re_right = list(filter(lambda x: x['signel'] == 'pass',
                           list(filter(lambda y: y['expected_command'] not in awake_command, data))))

    re_error = list(filter(lambda x: x['signel'] == 'fail',
                           list(filter(lambda y: y['expected_command'] not in awake_command, data))))

    re_null = count_times - len(re_right) - len(re_error)

    count_time = count_times + times

    count_rate = round((len(ak_right) + len(re_right)) / count_time * 100, 2)  # 总识别率
    count_false = round((len(ak_error) + len(re_error)) / count_time * 100, 2)  # 总错误率
    count_loss = round((ak_null + re_null) / count_time * 100, 2)  # 总未识别率

    try:
        awake_rate = round(len(ak_right) / times * 100, 2)  # 唤醒率
        awake_false = round(len(ak_error) / times * 100, 2)  # 错误率
        awake_loss = round(ak_null / times * 100, 2)  # 未识别率
    except:
        print(lines)
        print('测试日志无唤醒词！')
        awake_rate = 0
        awake_false = 0
        awake_loss = 0

    try:
        recognition_rate = round(len(re_right) / count_times * 100, 2)  # 识别率
        false_recognition = round(len(re_error) / count_times * 100, 2)  # 错误率
        un_recognition_rate = round(re_null / count_times * 100, 2)  # 未识别率
    except:
        print(lines)
        print('测试日志只有唤醒词！')
        recognition_rate = 0
        false_recognition = 0
        un_recognition_rate = 0

    # 计算每个命令词的次数
    pass_rate = list(map(lambda y: y['expected_command'],
                         list(filter(lambda x: x['signel'] == 'pass', data))))

    error_rate = list(map(lambda y: y['expected_command'],
                          list(filter(lambda x: x['signel'] == 'fail', data))))

    pass_list = list(map(pass_rate.count, all_command))
    error_list = list(map(error_rate.count, all_command))
    _count_list = list(map(count_command.count, all_command))

    loss_list = list(map(lambda x:
                         _count_list[x] - pass_list[x] - error_list[x], range(len(pass_list))))

    # 计算每个命令词的正确率, 误识率， 未识别率
    every_command_pass = list(map(lambda x: round(pass_list[x] / time_list[x] * 100, 2),
                                  range(len(pass_list))))
    every_command_fail = list(map(lambda x: round(error_list[x] / time_list[x] * 100, 2),
                                  range(len(error_list))))
    every_command_lost = list(map(lambda x: round(loss_list[x] / time_list[x] * 100, 2),
                                  range(len(loss_list))))

    return (
        awake_rate, awake_false, awake_loss, recognition_rate, false_recognition,
        un_recognition_rate, count_rate, count_false, count_loss,
        len(ak_right) + len(re_right), len(ak_right), len(re_right),
        len(ak_error) + len(re_error), len(ak_error), len(re_error),
        ak_null + re_null, ak_null, re_null,
        pass_list, error_list, loss_list,
        every_command_pass, every_command_fail, every_command_lost,
        _count_list
    )


@run_time()
def main(test_result_path, base_path, n: int = 1):
    """ 如果存在以前的excel结果，自动删除 """
    try:
        if os.path.exists(test_result_path):
            os.remove(test_result_path)
            print('已成功删除历史测试结果！')
    except Exception as e:
        print(f'历史测试结果xlsx不存在!-{e}')

    # 串口打印日志
    mic_log = get_mic_log(base_path)

    # 获取MIC.log文件内容 type list
    data = read_rs_trip_data(mic_log)  # 串口打印日志路径

    # 获取语音播放需要的命令词次数
    need_command = get_every_command_times(data)

    # 命令词数, 每个命令次出现的次数-list
    all_word, times = get_all_command_times(need_command)

    # 获取全部播放的音频
    require_wav = get_all_broadcast_wav(data)

    # 获取slaver_board的内容-list
    slaver_log = get_slaver_board_log(base_path)

    all_count_data = read_rs_trip_data(slaver_log)

    # 获取需要的日志内容
    all_res_list = get_res_count_data(all_count_data)

    # 获取MIC.log音频开始播放的时间
    all_start_time = get_start_time_list(data, all_res_list)

    """" *** 唤醒词个数 n，唤醒词语，唤醒词循环的次数，其他命令词次数*** """
    awake_command = all_word[:n]
    __wake_time = sum(times[:n])
    else_command_count_time = len(all_start_time) - __wake_time

    # pandas 将数据写入excel
    print(lines)
    print('Testing......')
    print(lines)

    all_wav = {
        'commands': need_command,
        'wav_name': require_wav,
        'start_time': all_start_time
    }

    (number, wav_name, start_time, expected_command,
     reback_time, reback_command, signel) = operation(all_res_list,
                                                      all_start_time,
                                                      require_wav,
                                                      need_command
                                                      )

    # case_id = list(range(1, len(all_res_list) + 1))
    have_res = {
        'case_id': number,
        'wav_name': wav_name,
        'start_time': start_time,
        'expected_command': expected_command,
        'reback_time': reback_time,
        'reback_command': reback_command,
        'signel': signel
    }

    res_list = [all_wav, have_res]
    sheet_1 = ['all_com_wav', 'pass_fail_info']
    try:
        to_excel.excel_add_sheet(res_list, sheet_1)
    except:
        print(lines)
        print('第一次写入测试excel的数据矩阵长度不等，无法成功写入......')
        sys.exit()

    # pandas 写入未识别的数据
    second_data = to_excel.get_data()
    lost_time = get_lost_wav_start_time(all_start_time, second_data)
    all_lost_wav = get_lost_wav(require_wav, second_data)
    all_command_list = get_lost_command(all_lost_wav, require_wav, need_command)
    lost_wav_command = {
        'wav_name': all_lost_wav,
        'start_time': lost_time,
        'commands': all_command_list,
        'signel': ['loss'] * len(all_lost_wav)
    }

    # 读取第一次写入excel结果
    sheet2_data = to_excel.read_data()

    (awake_rate, awake_false, awake_loss,
     recognition_rate, false_recognition, un_recognition_rate,
     _count_rate, _count_false, _count_loss,
     count_num, ak_num, re_num,
     count_err_num, count_ak_err, count_re_err,
     count_nu, ak_nu, re_nu,
     _pass_list, _error_list, _loss_list,
     every_command_pass_rate, every_command_fail_rate, every_command_lost_rate,
     __count_list) = recognize_rate(sheet2_data,
                                    else_command_count_time,
                                    __wake_time,
                                    awake_command,
                                    times,
                                    all_word,
                                    need_command
                                    )

    rate_dict = {
        '命令词': ['总命令词', '唤醒词', '其他命令词'] + all_word,
        '命令词总数': [len(need_command), __wake_time, else_command_count_time] + __count_list,
        '识别总数': [count_num, ak_num, re_num] + _pass_list,
        '误识数': [count_err_num, count_ak_err, count_re_err] + _error_list,
        '未识别数': [count_nu, ak_nu, re_nu] + _loss_list,
        '识别率(%)': [_count_rate, awake_rate, recognition_rate] + every_command_pass_rate,
        '误识率(%)': [_count_false, awake_false, false_recognition] + every_command_fail_rate,
        '未识别率(%)': [_count_loss, awake_loss, un_recognition_rate] + every_command_lost_rate
    }

    wav_list = [lost_wav_command, rate_dict]
    sheet_2 = ['loss_info', 'com_rate']
    try:
        to_excel.excel_add_sheet(wav_list, sheet_2)
    except:
        print(lines)
        print('第二次写入测试excel的数据矩阵长度不等，无法成功写入......')
        sys.exit()

    # 修改excel单元格显示问题
    to_excel.modify_excel()

    print(lines)
    print('Test finished !!!')
    print(lines)


if __name__ == "__main__":
    lines = '-----------------' * 2
    # 存放日志和测试结果的位置
    __base_path = 'D:\\'
    __test_result_path = os.path.join(__base_path, 'test_result.xlsx')
    # 初始化PandasManual
    to_excel = PandasManual(__test_result_path)

    __n = 3  # 传入唤醒词个数, 默认为 1，修改时只需改动 __n 即可
    main(__test_result_path, __base_path, n=__n)
    print(lines)
    print('测试结果在 {}'.format(__test_result_path))

