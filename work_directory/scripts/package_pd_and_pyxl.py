# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/9/27 12:06
# @Author  : Xiang Yu
# @File    : package_pd_and_pyxl.py
# @Software: PyCharm
# @Company : BEIJING INTENGINE


from openpyxl import load_workbook
from scripts.pandas_manual import PandasManual
from scripts.excel_manual import ExcelManual


class OperateExcel(PandasManual, ExcelManual):
    """ 二次封装，继承pandas和openpyxl """

    def __new__(cls, *args, **kwargs):
        """ 单例模式，重写init """
        return super().__new__(cls)

    def __init__(self, file_path):
        super(OperateExcel, self).__init__(file_path)
        self.file_path = file_path
        self.sheet = self.sheet_name

    def pd_write(self, data, sheet):
        """ pandas 写入excel方法 """
        super(OperateExcel, self).excel_add_sheet(data, sheet)
        # self.excel_add_sheet(data, sheet)

    @property
    def pd_read(self):
        """ pandas读取excel数据 """
        return super(OperateExcel, self).get_data

    def open_py_write(self, *args):
        """
        the_row, command, total, pass_num, fail_num, lost_num, pass_rate, fail_rate, lost_rate
        """
        super(OperateExcel, self).one_write_data(*args)

    @property
    def open_py_read(self):
        """
        read excel file
        :return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet[1]]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = []
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list

