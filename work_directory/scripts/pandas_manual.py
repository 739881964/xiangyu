# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/9/23 10:07
# @Author  : Xiang Yu
# @File    : pandas_manual.py
# @Software: PyCharm
# @Company : BEIJING INTENGINE

__all__ = ['FixExcel', 'ModifyExcelFormat', 'PandasManual']


import numpy as np

import os
import sys
import re
from datetime import datetime, timedelta
# sys.path.append('../')
# cur_path = os.path.abspath(os.path.dirname(__file__))
# root_path = os.path.split(cur_path)[0]
# sys.path.append(root_path)

import pandas as pd
from openpyxl import load_workbook, Workbook
from scripts.text_manual import run_time
from scripts.log_manual import log
from openpyxl.utils import get_column_letter
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side, Protection


class FixExcel:

    def __init__(self, file_path):
        self.file_path = file_path

    def modify_excel(self):
        wb = load_workbook(self.file_path)
        # 获取excel所有表单
        sheet_names = wb.sheetnames

        for k in range(len(sheet_names)):
            one_sheet = wb[sheet_names[k]]
            # 水平居中，垂直居中
            # one_sheet.alignment = Alignment(horizontal='center', vertical='center')
            # 获取每一列的内容的最大宽度
            m = 0
            # 每列
            col_width = list()
            for col in one_sheet.columns:
                # 每行
                for j in range(len(col)):
                    if j == 0:
                        # 数组增加一个元素
                        col_width.append(len(str(col[j].value)))
                    else:
                        # 获得每列中的内容的最大宽度
                        if col_width[m] < len(str(col[j].value)):
                            col_width[m] = len(str(col[j].value))
                m = m + 1

            # 设置列宽
            for m in range(len(col_width)):
                # 根据列的数字返回字母
                col_letter = get_column_letter(m + 1)
                # 当宽度大于100，宽度设置为100
                if col_width[m] > 80:
                    one_sheet.column_dimensions[col_letter].width = 80
                # 只有当宽度大于10，才设置列宽
                elif col_width[m] > 4:
                    one_sheet.column_dimensions[col_letter].width = col_width[m] + 4

        wb.save(filename=self.file_path)


class ModifyExcelFormat:

    def __init__(self, file_path):
        self.file_path = file_path

    @staticmethod
    def num2_column(num):
        """
        将列数转成列名对应单元格
        """
        interval = ord('Z') - ord('A')
        tmp = ''
        multiple = num // interval
        remainder = num % interval

        while multiple > 0:
            if multiple > 25:
                tmp += 'A'
            else:
                tmp += chr(64 + multiple)
            multiple = multiple // interval
        tmp += chr(64 + remainder)

        return tmp

    def func_openpyxl_modify_excel(self, df_list: list, no_list: list):
        """
        对Excel格式进行设置
        """
        wb = load_workbook(self.file_path)
        ws_list = wb.sheetnames

        border = Border(left=Side(style='thin', color='FF000000'),
                        right=Side(style='thin', color='FF000000'),
                        top=Side(style='thin', color='FF000000'),
                        bottom=Side(style='thin', color='FF000000'),
                        diagonal=Side(style='thin', color='FF000000'),
                        diagonal_direction=0, outline=Side(style='medium', color='FF000000'),
                        vertical=Side(style='thin', color='FF000000'),
                        horizontal=Side(style='thin', color='FF000000')
                        )

        # 循环表单
        for i in range(len(ws_list)):
            ws = wb[ws_list[i]]

            # 关闭默认灰色网格线
            ws.sheet_view.showGridLines = False

            # 第一行行高设置为22
            ws.row_dimensions[1].height = 22

            df = df_list[i]

            # 作为判断依据的列ID
            col_no = no_list[i]

            # 设置单元格边框
            for i in ws['A1:{}{}'.format(ModifyExcelFormat.num2_column(len(df.columns)), len(df) + 1)]:
                for j in i:
                    j.border = border

            # 对单元格进行填充
            fill_heading = PatternFill('solid', fgColor='BFBFBF')  # 灰色
            fill = PatternFill('solid', fgColor='FF9999')  # 亮粉

            for i in range(2, len(df) + 2):
                if ws.cell(row=i, column=col_no).value > 0:
                    for j in range(1, len(df.columns) + 1):
                        ws.cell(row=i, column=j).fill = fill_heading
                    ws.cell(row=i, column=col_no).fill = fill

            # ws水平居中，垂直居中
            for i in range(len(df)):  # 行
                for j in range(len(df.columns)):  # 列
                    cell = ws.cell(row=i + 2, column=j + 1)
                    cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平居中，垂直居中

            # ws自动设置列宽
            df_len = df.apply(lambda x: [(len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i)) for i in x], axis=0)
            df_len_max = df_len.apply(lambda x: max(x), axis=0)
            for i in df.columns:
                # 列的字母
                j = list(df.columns)
                column_letter = [chr(j.index(i) + 65) if j.index(i) <= 25 else 'A' + chr(j.index(i) - 26 + 65)][0]

                # 列的宽度
                columns_length = (len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i))
                data_max_length = df_len_max[i]
                column_width = [data_max_length if columns_length < data_max_length else columns_length][0]
                column_width = [column_width if column_width <= 50 else 50][0] + 3  # 列宽不能超过50

                # 更改列的宽度
                ws.column_dimensions['{}'.format(column_letter)].width = column_width

        wb.save(filename=self.file_path)


class PandasManual(FixExcel):
    """ pandas operate excel """

    def __new__(cls, *args, **kwargs):
        """ only create one object address """
        if not hasattr(PandasManual, '_instance'):
            cls._instance = super(PandasManual, cls).__new__(cls)

            return cls._instance

    def __init__(self, file_path):
        super().__init__(file_path)
        # self.file_path = file_path
        self.sheet_name = []

    def read_csv(self, sheet: 'data -> str' = None, encoding: 'format type' = 'gbk') -> object:
        """ read csv file content """
        data = None
        try:
            if sheet:
                data = pd.read_csv(self.file_path, sheet_name=sheet, encoding=encoding)
            else:
                data = pd.read_csv(self.file_path, encoding=encoding)
        except Exception as e:
            log.error(e)
            print(e)
            # raise e

        return data

    def get_data(self, sheet=None):
        """ read excel whole data by pandas"""
        if sheet:
            data = pd.read_excel(self.file_path, sheet_name=sheet)
        else:
            data = pd.read_excel(self.file_path, sheet_name=self.sheet_name[1])

        return data

    def read_data(self, sheet=''):
        """
        read excel file by openpyxl
        return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = wb[self.sheet_name[1]]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = list()
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list

    def open_write_data(self, the_row, wav_name1, start_time, wav_name2, end_time, spend_time, sheet_name=''):  # , expected_command, reback_time, reback_command, signel):
        """
        写入数据到excel,行
        """
        wb = load_workbook(self.file_path)
        if sheet_name:
            sheet = wb[sheet_name]
        else:
            sheet = wb.active

        if isinstance(the_row, int) and 2 <= the_row <= sheet.max_row:
            sheet.cell(the_row, 2).value = wav_name1  # config file's excel col
            sheet.cell(the_row, 3).value = start_time
            sheet.cell(the_row, 4).value = wav_name2  # config file's excel col
            sheet.cell(the_row, 5).value = end_time
            sheet.cell(the_row, 6).value = spend_time
            # sheet.cell(the_row, config.get_int('col', 'expected_command')).value = expected_command
            # sheet.cell(the_row, config.get_int('col', 'reback_time')).value = reback_time
            # sheet.cell(the_row, config.get_int('col', 'reback_command')).value = reback_command
            # sheet.cell(the_row, config.get_int('col', 'signel')).value = signel

            wb.save(self.file_path)
        else:
            print("unknow the_row")

    # @run_time(2)
    def write_data(self, rows: list, columns: list, res: list, sheet=None):
        """
        write in excel depend on row by openpyxl
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = Workbook().create_sheet('pass_fail_info')

        for row in rows:
            if isinstance(row, int) and 2 <= row <= sheet.max_row:
                for col in columns:
                    try:
                        sheet.cell(row, col).value = res[col]
                        wb.save(self.file_path)
                    except Exception as e:
                        log.error(e)
                        print("write in Excel failed")
                        # raise e
            else:
                print('row is not exist')

    # @run_time(2)
    def excel_add_sheet(self, data: list, sheet: list, _sheet_name='首页', header=None, index=None):
        """ not kill before operation when write after depend on column by pandas"""
        # if not exists excel txt, then create it
        try:
            if not os.path.exists(self.file_path):
                df = pd.DataFrame(['测试结果已生成!'])
                df.to_excel(self.file_path, sheet_name=_sheet_name, header=header, index=index)
                print('创建测试结果文件xlsx成功！')
        except (FileExistsError, Exception) as e:
            print(f'xlsx测试结果文件已存在，自动创建失败-{e}')
            raise e

        with pd.ExcelWriter(self.file_path) as excel_writer:
            book = load_workbook(excel_writer.path)
            excel_writer.book = book

            for i in range(len(data)):
                self.sheet_name.append(sheet[i])
                df = pd.DataFrame(data[i])
                df.to_excel(excel_writer=excel_writer, sheet_name=sheet[i], index=False)

            excel_writer.save()
        # excel_writer.close()


def get_time(data):
    """return wav datetime"""
    return datetime.strptime(data, "%Y-%m-%d %H:%M:%S")


def get_str_time(data):
    """get str-list"""
    return data.strftime('%Y-%m-%d %H:%M:%S')


if __name__ == '__main__':

    path = r'D:\\ret.xlsx'
    obj = PandasManual(path)
    data = obj.read_data(sheet='pass_fail_info')
    case_id = 0
    for i in range(1, len(data)):
        time = get_time(data[i]['start_time']) - get_time(data[i-1]['start_time'])
        h, m, s = str(time).split(':')
        time_len = int(h) * 3600 + int(m) * 60 + int(s)
        if time_len > 15:
            case_id += 1
            print(data[i]['wav_name'], data[i-1]['wav_name'])
            obj.open_write_data(case_id+1,
                                data[i-1]['wav_name'],
                                data[i-1]['start_time'],
                                data[i]['wav_name'],
                                data[i]['start_time'],
                                time_len,
                                sheet_name='ret'
                                )
        # if get_str_time(time)[-3:] > 15:
        #     print(data)

    # string = '5A 01 0A 07 00 04 00 00 00 00 00 52 0D'
    # print(list(string))
    # __base_path = 'D:\\'
    # __test_result_path = os.path.join(__base_path, 'test_result.xlsx')
    # panda = PandasManual(__test_result_path)
    # data = panda.get_data(sheet='com_rate')
    #
    # print(data['命令词'].tolist())
    # path = 'D:\\test_result.xlsx'
    # panda = PandasManual(path)
    # data = panda.get_data(sheet='all_com_wav')['wav_name'].tolist()
    # for i in data:
    #     if data.count(i) != 1:
    #         print(i)
    # all_file = os.listdir(r'C:\Users\xiangyu\Desktop\finally_list')
    # path = r'C:\Users\xiangyu\Desktop\多联41-60-new\1-20女'
    # for i in range(len(all_file)):
    #     # os.mkdir(os.path.join(path, all_file[i]))
    #     with open(os.path.join(path, all_file[i]), 'a+') as f:
    #         f.write('')
