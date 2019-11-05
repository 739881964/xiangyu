# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019-10-09 22:38
# @Author  : Xiang Yu
# @File    : excel_cell.py
# @Company : BEIJING-INTENGINE


from openpyxl import load_workbook
from openpyxl.styles import Color, Font, Alignment, PatternFill, Border, Side, Protection


# 将列数转成列名对应单元格
def num2_column(num):
    interval = ord('Z') - ord('A')
    tmp = ''
    multiple = num // interval
    remainder = num % interval
    while multiple > 0:
        if multiple > 25:
            tmp += 'A'
        else:
            tmp += chr(64 + multiple)
        multiple = multiple // interval
    tmp += chr(64 + remainder)
    return tmp


# 对Excel格式进行设置
def func_openpyxl_modify_excel(out_file, df_list, No_list):
    wb = load_workbook(out_file)
    ws_list = wb.sheetnames

    border = Border(left=Side(style='thin', color='FF000000'),
                    right=Side(style='thin', color='FF000000'),
                    top=Side(style='thin', color='FF000000'),
                    bottom=Side(style='thin', color='FF000000'),
                    diagonal=Side(style='thin', color='FF000000'),
                    diagonal_direction=0, outline=Side(style='medium', color='FF000000'),
                    vertical=Side(style='thin', color='FF000000'),
                    horizontal=Side(style='thin', color='FF000000'))

    for i in range(len(ws_list)):
        ws = wb[ws_list[i]]

        # 关闭默认灰色网格线
        ws.sheet_view.showGridLines = False

        # 第一行行高设置为22
        ws.row_dimensions[1].height = 22

        df = df_list[i]
        # 作为判断依据的列ID
        col_NO = No_list[i]
        # 设置单元格边框
        for i in ws['A1:{}{}'.format(num2_column(len(df.columns)), len(df) + 1)]:
            for j in i:
                j.border = border

        # 对单元格进行填充
        fill_heading = PatternFill('solid', fgColor='BFBFBF')  # 灰色
        fill = PatternFill('solid', fgColor='FF9999')  # 亮粉

        for i in range(2, len(df) + 2):
            if ws.cell(row=i, column=col_NO).value > 0:
                for j in range(1, len(df.columns) + 1):
                    ws.cell(row=i, column=j).fill = fill_heading
                ws.cell(row=i, column=col_NO).fill = fill

        # ws水平居中，垂直居中
        for i in range(len(df)):  # 行
            for j in range(len(df.columns)):  # 列
                cell = ws.cell(row=i + 2, column=j + 1)
                cell.alignment = Alignment(horizontal='center', vertical='center')  # 水平居中，垂直居中

        # ws自动设置列宽
        df_len = df.apply(lambda x: [(len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i)) for i in x], axis=0)
        df_len_max = df_len.apply(lambda x: max(x), axis=0)
        for i in df.columns:
            # 列的字母
            j = list(df.columns)
            column_letter = [chr(j.index(i) + 65) if j.index(i) <= 25 else 'A' + chr(j.index(i) - 26 + 65)][0]
            # 列的宽度
            columns_length = (len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i))
            data_max_length = df_len_max[i]
            column_width = [data_max_length if columns_length < data_max_length else columns_length][0]
            column_width = [column_width if column_width <= 50 else 50][0] + 3  # 列宽不能超过50
            # 更改列的宽度
            ws.column_dimensions['{}'.format(column_letter)].width = column_width

    wb.save(filename=out_file)

