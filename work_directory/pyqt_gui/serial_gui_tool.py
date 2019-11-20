# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019/11/20 10:08
# @Author  : Xiang Yu
# @File    : serial_gui_tool.py
# @Software: PyCharm
# @Company : BEIJING INTENGINE


__all__ = ['UiForm', 'PandasManual', 'PyQtSerial', '']


import sys
import serial
import serial.tools.list_ports
import pandas as pd

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QMessageBox, QPushButton, QFileDialog
from PyQt5.QtCore import QTimer
# from ui_demo_1 import Ui_Form
from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook


class UiForm(object):

    def __init__(self):
        pass

    def setup_ui(self, Form):
        Form.setObjectName("Form")
        Form.resize(707, 520)
        self.form_group_box = QtWidgets.QGroupBox(Form)
        self.form_group_box.setGeometry(QtCore.QRect(20, 20, 167, 301))
        self.form_group_box.setObjectName("form_group_box")
        self.formLayout = QtWidgets.QFormLayout(self.form_group_box)
        self.formLayout.setContentsMargins(10, 10, 10, 10)
        self.formLayout.setSpacing(10)
        self.formLayout.setObjectName("formLayout")
        self.s1__lb_1 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_1.setObjectName("s1__lb_1")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.s1__lb_1)
        self.s1__box_1 = QtWidgets.QPushButton(self.form_group_box)
        self.s1__box_1.setAutoRepeatInterval(100)
        self.s1__box_1.setDefault(True)
        self.s1__box_1.setObjectName("s1__box_1")
        self.formLayout.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.s1__box_1)
        self.s1__lb_2 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_2.setObjectName("s1__lb_2")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.s1__lb_2)
        self.s1__box_2 = QtWidgets.QComboBox(self.form_group_box)
        self.s1__box_2.setObjectName("s1__box_2")
        self.formLayout.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.s1__box_2)
        self.s1__lb_3 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_3.setObjectName("s1__lb_3")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.LabelRole, self.s1__lb_3)
        self.s1__box_3 = QtWidgets.QComboBox(self.form_group_box)
        self.s1__box_3.setObjectName("s1__box_3")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.s1__box_3.addItem("")
        self.formLayout.setWidget(3, QtWidgets.QFormLayout.FieldRole, self.s1__box_3)
        self.s1__lb_4 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_4.setObjectName("s1__lb_4")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.LabelRole, self.s1__lb_4)
        self.s1__box_4 = QtWidgets.QComboBox(self.form_group_box)
        self.s1__box_4.setObjectName("s1__box_4")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.s1__box_4.addItem("")
        self.formLayout.setWidget(4, QtWidgets.QFormLayout.FieldRole, self.s1__box_4)
        self.s1__lb_5 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_5.setObjectName("s1__lb_5")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.LabelRole, self.s1__lb_5)
        self.s1__box_5 = QtWidgets.QComboBox(self.form_group_box)
        self.s1__box_5.setObjectName("s1__box_5")
        self.s1__box_5.addItem("")
        self.formLayout.setWidget(5, QtWidgets.QFormLayout.FieldRole, self.s1__box_5)
        self.open_button = QtWidgets.QPushButton(self.form_group_box)
        self.open_button.setObjectName("open_button")
        self.formLayout.setWidget(7, QtWidgets.QFormLayout.SpanningRole, self.open_button)
        self.close_button = QtWidgets.QPushButton(self.form_group_box)
        self.close_button.setObjectName("close_button")
        self.formLayout.setWidget(8, QtWidgets.QFormLayout.SpanningRole, self.close_button)
        self.s1__lb_6 = QtWidgets.QLabel(self.form_group_box)
        self.s1__lb_6.setObjectName("s1__lb_6")
        self.formLayout.setWidget(6, QtWidgets.QFormLayout.LabelRole, self.s1__lb_6)
        self.s1__box_6 = QtWidgets.QComboBox(self.form_group_box)
        self.s1__box_6.setObjectName("s1__box_6")
        self.s1__box_6.addItem("")
        self.formLayout.setWidget(6, QtWidgets.QFormLayout.FieldRole, self.s1__box_6)
        self.state_label = QtWidgets.QLabel(self.form_group_box)
        self.state_label.setText("")
        self.state_label.setTextFormat(QtCore.Qt.AutoText)
        self.state_label.setScaledContents(True)
        self.state_label.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.state_label.setObjectName("state_label")
        self.formLayout.setWidget(2, QtWidgets.QFormLayout.SpanningRole, self.state_label)
        self.verticalGroupBox = QtWidgets.QGroupBox(Form)
        self.verticalGroupBox.setGeometry(QtCore.QRect(210, 20, 401, 241))
        self.verticalGroupBox.setObjectName("verticalGroupBox")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalGroupBox)
        self.verticalLayout.setContentsMargins(10, 10, 10, 10)
        self.verticalLayout.setObjectName("verticalLayout")
        self.s2__receive_text = QtWidgets.QTextBrowser(self.verticalGroupBox)
        self.s2__receive_text.setObjectName("s2__receive_text")
        self.verticalLayout.addWidget(self.s2__receive_text)
        self.verticalGroupBox_2 = QtWidgets.QGroupBox(Form)
        self.verticalGroupBox_2.setGeometry(QtCore.QRect(210, 280, 401, 101))
        self.verticalGroupBox_2.setObjectName("verticalGroupBox_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalGroupBox_2)
        self.verticalLayout_2.setContentsMargins(10, 10, 10, 10)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.s3__send_text = QtWidgets.QTextEdit(self.verticalGroupBox_2)
        self.s3__send_text.setObjectName("s3__send_text")
        self.verticalLayout_2.addWidget(self.s3__send_text)
        self.s3__send_button = QtWidgets.QPushButton(Form)
        self.s3__send_button.setGeometry(QtCore.QRect(620, 310, 61, 31))
        self.s3__send_button.setObjectName("s3__send_button")
        self.s3__clear_button = QtWidgets.QPushButton(Form)
        self.s3__clear_button.setGeometry(QtCore.QRect(620, 350, 61, 31))
        self.s3__clear_button.setObjectName("s3__clear_button")
        self.form_group_box1 = QtWidgets.QGroupBox(Form)
        self.form_group_box1.setGeometry(QtCore.QRect(20, 340, 171, 101))
        self.form_group_box1.setObjectName("form_group_box1")
        self.formLayout_2 = QtWidgets.QFormLayout(self.form_group_box1)
        self.formLayout_2.setContentsMargins(10, 10, 10, 10)
        self.formLayout_2.setSpacing(10)
        self.formLayout_2.setObjectName("formLayout_2")
        self.label = QtWidgets.QLabel(self.form_group_box1)
        self.label.setObjectName("label")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.LabelRole, self.label)
        self.label_2 = QtWidgets.QLabel(self.form_group_box1)
        self.label_2.setObjectName("label_2")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.LabelRole, self.label_2)
        self.lineEdit = QtWidgets.QLineEdit(self.form_group_box1)
        self.lineEdit.setObjectName("lineEdit")
        self.formLayout_2.setWidget(0, QtWidgets.QFormLayout.FieldRole, self.lineEdit)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.form_group_box1)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.formLayout_2.setWidget(1, QtWidgets.QFormLayout.FieldRole, self.lineEdit_2)
        self.hex_send = QtWidgets.QCheckBox(Form)
        self.hex_send.setGeometry(QtCore.QRect(620, 280, 71, 16))
        self.hex_send.setObjectName("hex_send")
        self.hex_receive = QtWidgets.QCheckBox(Form)
        self.hex_receive.setGeometry(QtCore.QRect(620, 40, 71, 16))
        self.hex_receive.setObjectName("hex_receive")
        self.s2__clear_button = QtWidgets.QPushButton(Form)
        self.s2__clear_button.setGeometry(QtCore.QRect(620, 80, 61, 31))
        self.s2__clear_button.setObjectName("s2__clear_button")
        self.timer_send_cb = QtWidgets.QCheckBox(Form)
        self.timer_send_cb.setGeometry(QtCore.QRect(260, 390, 71, 16))
        self.timer_send_cb.setObjectName("timer_send_cb")
        self.lineEdit_3 = QtWidgets.QLineEdit(Form)
        self.lineEdit_3.setGeometry(QtCore.QRect(350, 390, 61, 20))
        self.lineEdit_3.setAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignTrailing | QtCore.Qt.AlignVCenter)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.dw = QtWidgets.QLabel(Form)
        self.dw.setGeometry(QtCore.QRect(420, 390, 54, 20))
        self.dw.setObjectName("dw")
        self.verticalGroupBox.raise_()
        self.verticalGroupBox_2.raise_()
        self.form_group_box.raise_()
        self.s3__send_button.raise_()
        self.s3__clear_button.raise_()
        self.form_group_box.raise_()
        self.hex_send.raise_()
        self.hex_receive.raise_()
        self.s2__clear_button.raise_()
        self.timer_send_cb.raise_()
        self.lineEdit_3.raise_()
        self.dw.raise_()

        self.re_translate_ui(Form)
        QtCore.QMetaObject.connectSlotsByName(Form)

    def re_translate_ui(self, Form):
        _translate = QtCore.QCoreApplication.translate
        Form.setWindowTitle(_translate("Form", "Form"))
        self.form_group_box.setTitle(_translate("Form", "串口设置"))
        self.s1__lb_1.setText(_translate("Form", "串口检测："))
        self.s1__box_1.setText(_translate("Form", "检测串口"))
        self.s1__lb_2.setText(_translate("Form", "串口选择："))
        self.s1__lb_3.setText(_translate("Form", "波特率："))
        self.s1__box_3.setItemText(0, _translate("Form", "115200"))
        self.s1__box_3.setItemText(1, _translate("Form", "2400"))
        self.s1__box_3.setItemText(2, _translate("Form", "4800"))
        self.s1__box_3.setItemText(3, _translate("Form", "9600"))
        self.s1__box_3.setItemText(4, _translate("Form", "14400"))
        self.s1__box_3.setItemText(5, _translate("Form", "19200"))
        self.s1__box_3.setItemText(6, _translate("Form", "38400"))
        self.s1__box_3.setItemText(7, _translate("Form", "57600"))
        self.s1__box_3.setItemText(8, _translate("Form", "76800"))
        self.s1__box_3.setItemText(9, _translate("Form", "12800"))
        self.s1__box_3.setItemText(10, _translate("Form", "230400"))
        self.s1__box_3.setItemText(11, _translate("Form", "460800"))
        self.s1__lb_4.setText(_translate("Form", "数据位："))
        self.s1__box_4.setItemText(0, _translate("Form", "8"))
        self.s1__box_4.setItemText(1, _translate("Form", "7"))
        self.s1__box_4.setItemText(2, _translate("Form", "6"))
        self.s1__box_4.setItemText(3, _translate("Form", "5"))
        self.s1__lb_5.setText(_translate("Form", "校验位："))
        self.s1__box_5.setItemText(0, _translate("Form", "N"))
        self.open_button.setText(_translate("Form", "打开串口"))
        self.close_button.setText(_translate("Form", "关闭串口"))
        self.s1__lb_6.setText(_translate("Form", "停止位："))
        self.s1__box_6.setItemText(0, _translate("Form", "1"))
        self.verticalGroupBox.setTitle(_translate("Form", "接受区"))
        self.verticalGroupBox_2.setTitle(_translate("Form", "发送区"))
        self.s3__send_text.setHtml(_translate("Form", "<!DOCTYPE HTML PUBLIC \"-//W3C//DTD HTML 4.0//EN\" \"http://www.w3.org/TR/REC-html40/strict.dtd\">\n"
                                                                                "<html><head><meta name=\"qrichtext\" content=\"1\" /><style type=\"text/css\">\n"
                                                                                "p, li { white-space: pre-wrap; }\n"
                                                                                "</style></head><body style=\" font-family:\'SimSun\'; font-size:9pt; font-weight:400; font-style:normal;\">\n"
                                                                                "<p style=\" margin-top:0px; margin-bottom:0px; margin-left:0px; margin-right:0px; -qt-block-indent:0; text-indent:0px;\">123456</p></body></html>"
                                              )
                                   )
        self.s3__send_button.setText(_translate("Form", "发送"))
        self.s3__clear_button.setText(_translate("Form", "清除"))
        self.form_group_box1.setTitle(_translate("Form", "串口状态"))
        self.label.setText(_translate("Form", "已接收："))
        self.label_2.setText(_translate("Form", "已发送："))
        self.hex_send.setText(_translate("Form", "Hex发送"))
        self.hex_receive.setText(_translate("Form", "Hex接收"))
        self.s2__clear_button.setText(_translate("Form", "清除"))
        self.timer_send_cb.setText(_translate("Form", "定时发送"))
        self.lineEdit_3.setText(_translate("Form", "1000"))
        self.dw.setText(_translate("Form", "ms/次"))


class PandasManual:
    """ pandas operate excel """

    def __new__(cls, *args, **kwargs):
        """ only create one object address """
        if not hasattr(PandasManual, '_instance'):
            cls._instance = super(PandasManual, cls).__new__(cls)

            return cls._instance

    def __init__(self, file_path):
        # super().__init__(file_path)
        self.file_path = file_path
        self.sheet_name = []

    def get_data(self, sheet=''):
        """ read excel whole data by pandas"""
        if sheet:
            data = pd.read_excel(self.file_path, sheet_name=sheet)
        else:
            data = pd.read_excel(self.file_path, sheet_name=self.sheet_name[1])

        return data

    def read_data(self, sheet='') -> list:
        """
        read excel file by openpyxl
        return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        if sheet:
            sheet = wb[sheet]
        else:
            sheet = wb[self.sheet_name[1]]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = list()
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list


class PyQtSerial(QtWidgets.QWidget, UiForm, PandasManual):

    def __init__(self):
        super(PyQtSerial, self).__init__()
        self.myButton = QtWidgets.QPushButton(self)
        self.myButton.setObjectName("my_button")
        self.myButton.setText("选择文件/*.xlsx")
        self.myButton.clicked.connect(self.select_file)
        self.myButton.move(20, 450)
        self.setup_ui(self)
        self.init()
        self.setWindowTitle("串口小助手")
        self.ser = serial.Serial()
        self.port_check()
        # 接收数据和发送数据数目置零
        self.data_num_received = 0
        self.lineEdit.setText(str(self.data_num_received))
        self.data_num_send = 0
        self.lineEdit_2.setText(str(self.data_num_send))

    # @property
    def select_file(self):
        # directory = QFileDialog.getExistingDirectory(self, "选取文件夹", "./")  # 起始路径
        # print(directory)

        # 设置文件扩展名过滤,注意用双分号间隔
        all_file_name, file_type = QFileDialog.getOpenFileName(self, "选取文件", "./", "Text Files (*.xlsx)")

        print(all_file_name)

        return all_file_name

    def get_file_name(self):
        """返回PandasManual 实例化对象"""
        path = self.select_file()
        obj = PandasManual(path)

        return obj

    def init(self):
        # 串口检测按钮
        self.s1__box_1.clicked.connect(self.port_check)

        # 串口信息显示
        self.s1__box_2.currentTextChanged.connect(self.port_imf)

        # 打开串口按钮
        self.open_button.clicked.connect(self.port_open)

        # 关闭串口按钮
        self.close_button.clicked.connect(self.port_close)

        # 发送数据按钮
        self.s3__send_button.clicked.connect(self.data_send)

        # 定时发送数据
        self.timer_send = QTimer()
        self.timer_send.timeout.connect(self.data_send)
        self.timer_send_cb.stateChanged.connect(self.data_send_timer)

        # 定时器接收数据
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.data_receive)

        # 清除发送窗口
        self.s3__clear_button.clicked.connect(self.send_data_clear)

        # 清除接收窗口
        self.s2__clear_button.clicked.connect(self.receive_data_clear)

    # 串口检测
    def port_check(self):
        # 检测所有存在的串口，将信息存储在字典中
        self.Com_Dict = {}
        port_list = list(serial.tools.list_ports.comports())
        self.s1__box_2.clear()
        for port in port_list:
            self.Com_Dict["%s" % port[0]] = "%s" % port[1]
            self.s1__box_2.addItem(port[0])
        if len(self.Com_Dict) == 0:
            self.state_label.setText(" 无串口")

    # 串口信息
    def port_imf(self):
        # 显示选定的串口的详细信息
        imf_s = self.s1__box_2.currentText()
        if imf_s != "":
            self.state_label.setText(self.Com_Dict[self.s1__box_2.currentText()])

    # 打开串口
    def port_open(self):
        self.ser.port = self.s1__box_2.currentText()
        self.ser.baudrate = int(self.s1__box_3.currentText())
        self.ser.bytesize = int(self.s1__box_4.currentText())
        self.ser.stopbits = int(self.s1__box_6.currentText())
        self.ser.parity = self.s1__box_5.currentText()

        try:
            self.ser.open()
        except:
            QMessageBox.critical(self, "Port Error", "此串口不能被打开！")
            return

        # 打开串口接收定时器，周期为2ms
        self.timer.start(2)

        if self.ser.isOpen():
            self.open_button.setEnabled(False)
            self.close_button.setEnabled(True)
            self.form_group_box1.setTitle("串口状态（已开启）")

    # 关闭串口
    def port_close(self):
        self.timer.stop()
        self.timer_send.stop()
        try:
            self.ser.close()
        except:
            pass
        self.open_button.setEnabled(True)
        self.close_button.setEnabled(False)
        self.lineEdit_3.setEnabled(True)
        # 接收数据和发送数据数目置零
        self.data_num_received = 0
        self.lineEdit.setText(str(self.data_num_received))
        self.data_num_sended = 0
        self.lineEdit_2.setText(str(self.data_num_sended))
        self.form_group_box1.setTitle("串口状态（已关闭）")

    # 发送数据
    def data_send(self):
        if self.ser.isOpen():
            input_s = self.s3__send_text.toPlainText()
            if input_s != "":
                # 非空字符串
                if self.hex_send.isChecked():
                    # hex发送
                    input_s = input_s.strip()
                    send_list = []
                    while input_s != '':
                        try:
                            num = int(input_s[0:2], 16)
                        except ValueError:
                            QMessageBox.critical(self, 'wrong data', '请输入十六进制数据，以空格分开!')
                            return None
                        input_s = input_s[2:].strip()
                        send_list.append(num)
                    input_s = bytes(send_list)
                else:
                    # ascii发送
                    input_s = (input_s + '\r\n').encode('utf-8')

                num = self.ser.write(input_s)
                self.data_num_sended += num
                self.lineEdit_2.setText(str(self.data_num_sended))
        else:
            pass

    # 接收数据
    def data_receive(self):
        data = self.obj.read_data()
        try:
            num = self.ser.inWaiting()
        except:
            self.port_close()
            return
        if num > 0:
            data = self.ser.read(num)
            num = len(data)
            # hex显示
            if self.hex_receive.checkState():
                out_s = ''
                for i in range(0, len(data)):
                    out_s = out_s + '{:02X}'.format(data[i]) + ' '
                for one_data in data:
                    if one_data['语音模块发送给主控MCU的UART数据'] == out_s:
                        self.s2__receive_text.insertPlainText('响应串口码: [{}] -> 对应的命令或状态：[{}]'.format(out_s, one_data['命令 或 状态']))
                    else:
                        self.s2__receive_text.insertPlainText('响应的串口码 [{}] 无法在excel中找到'.format(out_s))
            else:
                # 串口接收到的字符串为b'123',要转化成unicode字符串才能输出到窗口中去
                self.s2__receive_text.insertPlainText(data.decode('utf-8'))

            # 统计接收字符的数量
            self.data_num_received += num
            self.lineEdit.setText(str(self.data_num_received))

            # 获取到text光标
            text_cursor = self.s2__receive_text.textCursor()
            # 滚动到底部
            text_cursor.movePosition(text_cursor.End)
            # 设置光标到text中去
            self.s2__receive_text.setTextCursor(text_cursor)
        else:
            pass

    # 定时发送数据
    def data_send_timer(self):
        if self.timer_send_cb.isChecked():
            self.timer_send.start(int(self.lineEdit_3.text()))
            self.lineEdit_3.setEnabled(False)
        else:
            self.timer_send.stop()
            self.lineEdit_3.setEnabled(True)

    # 清除显示
    def send_data_clear(self):
        self.s3__send_text.setText("")

    def receive_data_clear(self):
        self.s2__receive_text.setText("")


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    my_show = PyQtSerial()
    my_show.show()
    sys.exit(app.exec_())

