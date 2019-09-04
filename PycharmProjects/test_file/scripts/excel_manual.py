# coding=utf-8
from openpyxl import *
from scripts.base_path import EXCEL_PATH
from scripts.conf_manual import config


class ExcelManual(object):

    def __init__(self, file_path, sheet_name):
        """
        初始化文件路径与sheet名称
        :param file_path: 文件路径
        :param sheet_name: sheet表单名称
        """
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_data(self):
        """
        读取excel文件
        :return: 内容为嵌套dict-list
        """
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = []
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list

    def write_data(self, the_row, actual_res, res=None):
        """
        写入数据到excel
        :param the_row: 写入的行
        :param actual_res: 实际结果
        :param res: pass或fail
        :return: None
        """
        wb = load_workbook(self.file_path)
        if self.sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]

        # sheet.cell(the_row, config.get_int('col', 'actual_res')).value = actual_res
        # wb.save(self.file_path)

        if isinstance(the_row, int) and 2 <= the_row <= sheet.max_row:
            sheet.cell(the_row, config.get_int('col', 'actual_res')).value = actual_res  # 配置文件中Excel列号
            sheet.cell(the_row, config.get_int('col', 'res')).value = res
            wb.save(self.file_path)
        else:
            print("unknow the_row")


if __name__ == '__main__':
    excel = ExcelManual(EXCEL_PATH, 'user_register')
    __data = excel.read_data()
    # excel.write_data(16, 1, 2, 3, 4)
    # for i in __data:
    #     print(i)
    # print(len(__data))
    print(__data)
