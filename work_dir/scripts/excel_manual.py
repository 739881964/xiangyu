# !/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2019-09-05 22:45
# @Author  : Yu xiang
# @File    : excel_manual.py
# @Company : BEIJING INTENGINE

from openpyxl import *
from scripts.base_path import EXCEL_PATH
from scripts.conf_manual import config


class ExcelManual(object):
    def __init__(self, file_path, sheet_name):
        """
        init path and sheet name
        :param file_path: file_path
        :param sheet_name: sheet_name
        """
        self.file_path = file_path
        self.sheet_name = sheet_name

    def read_data(self):
        """
        read excel file
        :return: content is dict-list
        """
        wb = load_workbook(self.file_path)
        sheet = wb[self.sheet_name]
        head_data = list(sheet.iter_rows(max_row=2, values_only=True))[0]

        data_list = []
        for data in list(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)):
            dic = dict(zip(head_data, data))
            data_list.append(dic)

        return data_list

    def write_data(self, the_row, wav_name, start_time, expected_command, reback_time, reback_command, signel):
        """
        write data to excel
        :param the_row: row
        :param actual_res: actual result
        :param res: pass or fail
        :return: None
        """
        wb = load_workbook(self.file_path)
        if self.sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]

        if isinstance(the_row, int) and 2 <= the_row <= sheet.max_row:
            sheet.cell(the_row, config.get_int('col', 'wav_name')).value = wav_name  # config file's excel col
            sheet.cell(the_row, config.get_int('col', 'start_time')).value = start_time
            sheet.cell(the_row, config.get_int('col', 'expected_command')).value = expected_command
            sheet.cell(the_row, config.get_int('col', 'reback_time')).value = reback_time
            sheet.cell(the_row, config.get_int('col', 'reback_command')).value = reback_command
            sheet.cell(the_row, config.get_int('col', 'signel')).value = signel

            wb.save(self.file_path)
        else:
            print("unknow the_row")

    def one_write_data(self, the_row, command, total, pass_num, fail_num, lost_num, pass_rate, fail_rate, lost_rate):
        """
         :write data to excel
         :param the_row: row
         :param actual_res: actual result
         :param res: pass or fail
         :return: None
         """
        wb = load_workbook(self.file_path)
        if self.sheet_name is None:
            sheet = wb.active
        else:
            sheet = wb[self.sheet_name]

        if isinstance(the_row, int) and 2 <= the_row <= sheet.max_row:
            sheet.cell(the_row, config.get_int('one_col', 'command')).value = command  # config file's excel col
            sheet.cell(the_row, config.get_int('one_col', 'total')).value = total
            sheet.cell(the_row, config.get_int('one_col', 'pass_num')).value = pass_num
            sheet.cell(the_row, config.get_int('one_col', 'fail_num')).value = fail_num
            sheet.cell(the_row, config.get_int('one_col', 'lost_num')).value = lost_num
            sheet.cell(the_row, config.get_int('one_col', 'pass_rate')).value = pass_rate
            sheet.cell(the_row, config.get_int('one_col', 'fail_rate')).value = fail_rate
            sheet.cell(the_row, config.get_int('one_col', 'lost_rate')).value = lost_rate

            wb.save(self.file_path)
        else:
            print("unknow the_row")


if __name__ == '__main__':
    excel = ExcelManual(EXCEL_PATH, 'user_register')
    __data = excel.read_data()
    # excel.write_data(16, 1, 2, 3, 4)
    # for i in __data:
    #     print(i)
    # print(len(__data))
    print(__data)
